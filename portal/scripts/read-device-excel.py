import base64
import csv
import io
import json
import re
import sys

sys.stdin.reconfigure(encoding="utf-8")
sys.stdout.reconfigure(encoding="utf-8")

payload = json.load(sys.stdin)
name = str(payload.get("name", "")).lower()
raw = base64.b64decode(payload.get("data", ""))

if name.endswith(".csv"):
    text = raw.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
elif name.endswith(".xlsx"):
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = []
    known_headers = {
        "设备资产编码", "资产编号", "资产编码", "固定资产编码", "资产卡片编号", "编号",
        "类型", "设备类别", "资产类别", "设备品牌型号", "品牌型号", "规格型号",
        "设备名称", "资产名称", "序列号", "使用人姓名", "使用人", "部门", "使用部门",
        "采购时间", "购入日期", "购买金额", "资产原值", "状态", "存放地点"
    }
    special_inventory = any(
        sheet.max_row >= 2
        and "姓名" in [str(x or "").strip() for x in next(sheet.iter_rows(values_only=True))]
        and "电脑" in [str(x or "").strip() for x in next(sheet.iter_rows(values_only=True))]
        for sheet in workbook.worksheets
        if sheet.title != "汇总"
    )

    def clean(value):
        return str(value or "").strip()

    def pieces(value):
        text = clean(value)
        if not text or text in {"/", "-", "无", "否"}:
            return []
        # 多台设备常用换行、带空格的斜杠或“1./2.”分隔；型号内部无空格斜杠仍保留。
        text = re.sub(r"(?<!^)\s+(?=\d+[.、．]\s*)", "\n", text)
        parts = [re.sub(r"^\s*\d+[.、．]\s*", "", x).strip() for x in re.split(r"[\r\n]+|\s+/\s+", text)]
        return [x for x in parts if x and x not in {"/", "-"}]

    def category_of(value):
        text = clean(value).lower()
        if "显示" in text or "屏" in text:
            return "显示器"
        if "一体" in text or "imac" in text:
            return "一体机"
        if "台式" in text or "主机" in text or "组装" in text:
            return "台式电脑"
        if "mini" in text:
            return "Mac Mini"
        if "手机" in text or "iphone" in text:
            return "手机"
        if "平板" in text or "ipad" in text:
            return "平板电脑"
        return "笔记本"

    def department_name(title):
        return re.split(r"[-【]", title, maxsplit=1)[0].strip()

    def append_asset(sheet, excel_row, person, employee_no, job, workstation, category, model, serial, notes="", status="出库", company_issued=""):
        code = clean(serial)
        model = clean(model)
        category = clean(category) or ("一体机" if "一体" in model else "笔记本")
        if not code or not model:
            return
        rows.append({
            "设备资产编码": code,
            "类型": category,
            "设备品牌型号": model,
            "使用人姓名": clean(person),
            "部门": department_name(sheet.title),
            "职位": clean(job),
            "具体地点": clean(workstation),
            "备注": clean(notes),
            "状态": status,
            "是否公司配发设备": clean(company_issued),
            "员工工号": clean(employee_no),
            "__sheet": sheet.title,
            "__row": excel_row,
        })

    # 新版公司盘点表采用两层表头，而且“汇总”与各部门工作表内容重复。
    # 这里跳过汇总页，从各部门页提取电脑和显示器，再转换为首版标准字段。
    if special_inventory:
        ignored = {"汇总", "报废电脑", "收回电脑", "新采购电脑", "维修"}
        for sheet in workbook.worksheets:
            if sheet.title in ignored:
                continue
            values = list(sheet.iter_rows(values_only=True))
            if len(values) < 3:
                continue
            top, sub = list(values[0]), list(values[1])
            top_text = [clean(x).replace("\n", "") for x in top]
            if "姓名" not in top_text or "电脑" not in top_text:
                continue
            def col(label):
                return top_text.index(label) if label in top_text else -1
            name_i, no_i, job_i, seat_i = col("姓名"), col("工号"), col("职位"), col("工位号")
            computer_i = col("电脑")
            display_i = col("显示屏")
            issued_columns = [i for i, value in enumerate(top_text) if value == "是否使用公司设备"]
            computer_issued_i = max((i for i in issued_columns if i < computer_i), default=-1)
            display_issued_i = max((i for i in issued_columns if i < display_i), default=-1) if display_i >= 0 else -1
            computer_end = display_i if display_i > computer_i else len(top)
            # 备注有时位于电脑与显示屏之间，不能当作设备字段。
            for i in range(computer_i + 1, computer_end):
                if top_text[i] in {"备注", "是否使用公司设备"}:
                    computer_end = i
                    break
            for excel_row, data in enumerate(values[2:], start=3):
                person = data[name_i] if name_i >= 0 and name_i < len(data) else None
                employee_no = data[no_i] if no_i >= 0 and no_i < len(data) else None
                job = data[job_i] if job_i >= 0 and job_i < len(data) else None
                workstation = data[seat_i] if seat_i >= 0 and seat_i < len(data) else None
                computer_issued = data[computer_issued_i] if computer_issued_i >= 0 and computer_issued_i < len(data) else ""
                display_issued = data[display_issued_i] if display_issued_i >= 0 and display_issued_i < len(data) else ""
                if not clean(person) or clean(person) == "姓名":
                    continue
                group = [(i, clean(sub[i]).replace("\n", "")) for i in range(max(0, computer_i), min(computer_end, len(sub)))]
                serial_i = next((i for i, h in reversed(group) if "SN" in h.upper() or "S/N" in h.upper()), -1)
                type_i = next((i for i, h in group if "类型" in h), -1)
                model_i = next((i for i, h in group if "型号" in h and i != serial_i), -1)
                serials = pieces(data[serial_i]) if serial_i >= 0 and serial_i < len(data) else []
                models = pieces(data[model_i]) if model_i >= 0 and model_i < len(data) else []
                types = pieces(data[type_i]) if type_i >= 0 and type_i < len(data) else []
                # “第一台无 SN、第二台有 SN”时，从末尾对齐，避免把第二台 SN 配给第一台型号。
                if serials and len(types) > len(serials):
                    types = types[-len(serials):]
                if serials and len(models) > len(serials):
                    models = models[-len(serials):]
                for index, serial in enumerate(serials):
                    type_text = types[index] if index < len(types) else (types[-1] if types else "")
                    model_text = models[index] if index < len(models) else (models[-1] if models else "")
                    category = category_of(f"{type_text} {model_text}")
                    generic = type_text in {"笔记本", "台式电脑", "一体机", "电脑", "Mac一体", "Mac Mini"}
                    model = model_text if generic else " ".join(x for x in [type_text, model_text] if x).strip()
                    append_asset(sheet, excel_row, person, employee_no, job, workstation, category, model or category, serial, company_issued=computer_issued)
                if display_i >= 0:
                    display_end = len(top)
                    for i in range(display_i + 1, len(top)):
                        if top_text[i] in {"是否离职", "是否归还设备", "更换设备", "报废", "备注"}:
                            display_end = i
                            break
                    display_group = [(i, clean(sub[i]).replace("\n", "")) for i in range(display_i, min(display_end, len(sub)))]
                    display_value_i = next((i for i, h in reversed(display_group) if "型号" in h or "SN" in h.upper()), -1)
                    if display_value_i >= 0 and display_value_i < len(data):
                        display_text = clean(data[display_value_i])
                        sn_match = re.search(r"SN\s*码?[:：]\s*([^\r\n]+)", display_text, flags=re.I)
                        if sn_match:
                            serial = sn_match.group(1).strip()
                            model = re.sub(r"[\r\n]*SN\s*码?[:：].*$", "", display_text, flags=re.I | re.S).strip() or "显示器"
                            append_asset(sheet, excel_row, person, employee_no, job, workstation, "显示器", model, serial, company_issued=display_issued)
                        else:
                            displays = pieces(display_text)
                            for display in displays:
                                pair = [x.strip() for x in re.split(r"\s*/\s*", display, maxsplit=1)]
                                if len(pair) == 2 and pair[0] and pair[1]:
                                    append_asset(sheet, excel_row, person, employee_no, job, workstation, "显示器", pair[0], pair[1], company_issued=display_issued)
                                else:
                                    append_asset(sheet, excel_row, person, employee_no, job, workstation, "显示器", "型号未记录", display, company_issued=display_issued)
    else:
      for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        best_index, best_score = 0, -1
        for index, candidate in enumerate(values[:50]):
            cells = {str(value or "").strip().replace("\n", "") for value in candidate}
            score = sum(1 for header in known_headers if header.replace("\n", "") in cells)
            if score > best_score:
                best_index, best_score = index, score
        headers = [str(x or "").strip() for x in values[best_index]]
        for excel_row, row in enumerate(values[best_index + 1:], start=best_index + 2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = {headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]}
            record["__sheet"] = sheet.title
            record["__row"] = excel_row
            rows.append(record)
else:
    raise ValueError("请将旧版 .xls 文件另存为 .xlsx 后再导入")

# 不依赖文件名或工作表名：最终按资产编码/SN 去重，避免汇总页和明细页重复。
unique = {}
for row in rows:
    code = re.sub(r"\s+", "", str(row.get("设备资产编码") or row.get("资产编号") or "")).upper()
    key = code or f"__row__{len(unique)}"
    old = unique.get(key)
    if old is None or sum(bool(v) for v in row.values()) > sum(bool(v) for v in old.values()):
        unique[key] = row
rows = list(unique.values())

print(json.dumps(rows, ensure_ascii=False, default=str))
